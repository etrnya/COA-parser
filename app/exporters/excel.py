import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.utils.logger import get_logger

logger = get_logger("ExcelExporter")

# Column ID to User-friendly Header Label Mapping
HEADER_LABELS = {
    "file_name": "檔案名稱",
    "product_raw": "產品名稱 (原始)",
    "product_std": "產品名稱 (標準化)",
    "brand_raw": "廠牌 (原始)",
    "brand_std": "廠牌 (標準化)",
    "batch_no": "生產批號",
    "expiry_raw": "有效期限 (原始)",
    "expiry_std": "有效期限 (標準化)",
    "purity_raw": "純度/含量 (原始)",
    "purity_std": "純度/含量 (標準化)",
    "amount_raw": "包裝容量 (原始)",
    "amount_std": "包裝容量 (標準化)",
    "cas_no_raw": "CAS Number (原始)",
    "cas_no_std": "CAS Number (標準化)",
    "mw_raw": "分子量 (原始)",
    "mw_std": "分子量 (標準化)",
    "storage_raw": "儲存條件 (原始)",
    "storage_std": "儲存條件 (標準化對應)"
}

import re

def sanitize_value(val):
    """Remove control characters that are invalid in Excel XML (except tab, newline, carriage return)."""
    if not isinstance(val, str):
        return val
    # Remove ASCII control characters (0-31) except 9 (tab), 10 (newline), 13 (carriage return)
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)

def export_to_excel(
    records: list, 
    output_path: str, 
    fields_order: list, 
    fields_visibility: dict,
    column_headers: dict = None
) -> bool:
    """
    Export verified COA records into a professionally styled Excel spreadsheet.
    
    Args:
        records: List of dicts representing verified COA tasks.
        output_path: Destination path for the .xlsx file.
        fields_order: Ordered list of column IDs.
        fields_visibility: Dict mapping column IDs to booleans (whether to include them).
        column_headers: Optional dictionary containing user-configured header names.
    """
    logger.info(f"Starting Excel export for {len(records)} records...")
    
    try:
        # 1. Flatten records into rows mapping to fields_order
        rows = []
        for rec in records:
            data = rec.get("extracted_data") or {}
            
            row = {"file_name": sanitize_value(rec.get("file_name", "Unknown"))}
            for field in fields_order:
                val = data.get(field)
                if val is None:
                    val = "N/A"
                row[field] = sanitize_value(val)
            rows.append(row)
            
        # 2. Filter out invisible columns
        # First column is always 'file_name'
        visible_columns = ["file_name"] + [f for f in fields_order if fields_visibility.get(f, True)]
        
        df = pd.DataFrame(rows)
        # Reorder and filter columns
        df = df[visible_columns]
        
        # Rename columns to user-friendly labels
        headers_map = column_headers if column_headers else HEADER_LABELS
        friendly_headers = {col: headers_map.get(col, HEADER_LABELS.get(col, col)) for col in visible_columns}
        df = df.rename(columns=friendly_headers)
        
        # 3. Write using Pandas with openpyxl engine
        df.to_excel(output_path, index=False, sheet_name="COA Summary")
        
        # 4. Inject Premium Styling via openpyxl
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Colors & Styling definitions
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # Format headers
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # Format data cells & apply thin borders
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border
                # Left align for text, center for dates/purity/batch
                if col_idx == 1 or "Name" in str(ws.cell(row=1, column=col_idx).value):
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        # Apply Auto-Filter
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        
        # Freeze top row
        ws.freeze_panes = "A2"
        
        # Auto-adjust column widths to prevent ### errors
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            # Add padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Save modifications
        wb.save(output_path)
        wb.close()
        logger.info(f"Excel file successfully generated and formatted at: {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        return False
